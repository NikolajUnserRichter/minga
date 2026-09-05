"""Excel-Upload für Stammdaten + Template-Download.

Endpoints:
- GET  /imports/template/{entity}   → XLSX mit Header-Zeile + Beispiel
- POST /imports/{entity}            → Multipart-Upload, parst Rows + upsertet

Unterstützte entities: customers, seeds, products, suppliers, locations
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional, Sequence
from uuid import UUID

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select

from app.api.deps import DBSession
from app.models.customer import Customer, CustomerType
from app.services.customer_service import next_customer_number
from app.models.seed import Seed, Supplier
from app.models.product import Product, ProductCategory
from app.models.inventory import InventoryLocation, LocationType
from app.models.unit import UnitOfMeasure
from app.models.enums import TaxRate
from app.models.order import Order, OrderLine, OrderStatus

router = APIRouter(prefix="/imports", tags=["Excel-Import"])

# Blattnamen im Template: Daten wird importiert, Beispiel nie.
DATA_SHEET = "Daten"
EXAMPLE_SHEET = "Beispiel"


# ---- Spalten-Definitionen je Entity ---------------------------------------

# Format: (column_header, attr_name, required, type_hint)
COLUMNS = {
    "customers": [
        ("name", "name", True, "str"),
        ("typ", "typ", True, "enum:GASTRO|HANDEL|GEWERBE|PRIVAT"),
        ("email", "email", False, "str"),
        ("telefon", "telefon", False, "str"),
        ("adresse", "adresse", False, "str"),
        ("ust_id", "ust_id", False, "str"),
        ("notizen", "notizen", False, "str"),
    ],
    "suppliers": [
        ("name", "name", True, "str"),
        ("email", "email", False, "str"),
        ("telefon", "telefon", False, "str"),
        ("adresse", "adresse", False, "str"),
        ("ust_id", "ust_id", False, "str"),
        ("product_group", "product_group", False, "enum:SAATGUT|SUBSTRAT|VERPACKUNG|ARBEITSMATERIAL|SONSTIGES"),
        ("is_organic", "is_organic", False, "bool"),
        ("bio_kontrollstelle", "bio_kontrollstelle", False, "str"),
        ("notizen", "notizen", False, "str"),
    ],
    "seeds": [
        ("name", "name", True, "str"),
        ("sorte", "sorte", False, "str"),
        ("lieferant", "lieferant", False, "str"),
        ("keimdauer_tage", "keimdauer_tage", True, "int"),
        ("wachstumsdauer_tage", "wachstumsdauer_tage", True, "int"),
        ("erntefenster_min_tage", "erntefenster_min_tage", True, "int"),
        ("erntefenster_optimal_tage", "erntefenster_optimal_tage", True, "int"),
        ("erntefenster_max_tage", "erntefenster_max_tage", True, "int"),
        ("ertrag_gramm_pro_tray", "ertrag_gramm_pro_tray", True, "decimal"),
        ("verlustquote_prozent", "verlustquote_prozent", False, "decimal"),
        ("saatgut_pro_einheit_gramm", "saatgut_pro_einheit_gramm", False, "decimal"),
        ("cooling_days", "cooling_days", False, "int"),
        ("cooling_shelf_life_days", "cooling_shelf_life_days", False, "int"),
        ("process_type", "process_type", False, "enum:STANDARD|PLATTE|PLATTE_STEINE"),
    ],
    "products": [
        ("sku", "sku", True, "str"),
        ("name", "name", True, "str"),
        # Sortenreine Artikel führen die Sorte als Zusatzzeile auf Lieferschein
        # und Rechnung — der Artikelname allein sagt sie nicht.
        ("sorte", "seed_variety", False, "str"),
        ("category", "category", True, "enum:MICROGREEN|SEED|PACKAGING|BUNDLE"),
        ("gtin", "gtin", False, "str"),
        ("old_article_number", "old_article_number", False, "str"),
        ("certification", "certification", False, "enum:BIO|KONVENTIONELL|TRANSITIONAL"),
        ("description", "description", False, "str"),
        ("base_price", "base_price", False, "decimal"),
        ("tax_rate", "tax_rate", False, "enum:REDUZIERT|STANDARD|STEUERFREI"),
        ("shelf_life_days", "shelf_life_days", False, "int"),
        # Pfandgebinde (Mehrwegtray, Pfandkiste): tax_rate gehört auf STANDARD,
        # Pfand ist kein Lebensmittelumsatz.
        ("pfand", "is_deposit", False, "bool"),
        ("pfandwert", "deposit_value", False, "decimal"),
    ],
    "locations": [
        ("code", "code", True, "str"),
        ("name", "name", True, "str"),
        ("location_type", "location_type", True, "enum:LAGER|KUEHLRAUM|REGAL|KEIMRAUM|VERSAND"),
        ("description", "description", False, "str"),
        ("temperature_min", "temperature_min", False, "decimal"),
        ("temperature_max", "temperature_max", False, "decimal"),
    ],
    "order_history": [
        ("bestell_nr_extern", "bestell_nr_extern", True, "str"),
        ("kunde", "kunde", True, "str"),
        ("bestelldatum", "bestelldatum", True, "date"),
        ("lieferdatum", "lieferdatum", True, "date"),
        ("produkt_sku", "produkt_sku", True, "str"),
        ("menge", "menge", True, "decimal"),
        ("einheit", "einheit", False, "str"),
        ("einzelpreis", "einzelpreis", True, "decimal"),
        ("status", "status", False, "enum:ENTWURF|BESTAETIGT|IN_PRODUKTION|GELIEFERT|FAKTURIERT|STORNIERT"),
        # JSON-Array für Variable-Bundle-Sorten:
        # z.B. [{"sku": "MG-SONNE", "quantity": 1}, {"sku": "MG-ERBSE", "quantity": 1}]
        # SKUs werden gegen die Produkt-Tabelle aufgelöst.
        ("bundle_selections", "bundle_selections", False, "str"),
    ],
    # Go-Live: historische/laufende Wachstumschargen (Sorte muss als Saatgut existieren)
    "grow_batches": [
        ("sorte", "sorte", True, "str"),
        ("aussaat_datum", "aussaat_datum", True, "date"),
        ("tray_anzahl", "tray_anzahl", True, "int"),
        ("status", "status", False, "enum:KEIMUNG|WACHSTUM|ERNTEREIF|GEERNTET|VERLUST"),
        ("charge_nummer", "charge_nummer", False, "str"),
        ("regal_position", "regal_position", False, "str"),
        ("ernte_datum", "ernte_datum", False, "date"),
        ("ernte_menge_stueck", "ernte_menge_stueck", False, "int"),
        ("ernte_menge_gramm", "ernte_menge_gramm", False, "decimal"),
        # Historien-Import (Warenfluss-Release): Idempotenzschlüssel,
        # Bestandswirkung und Ausschuss je Charge
        ("externe_chargennummer", "externe_chargennummer", False, "str"),
        ("saatgut_gramm", "saatgut_gramm", False, "decimal"),
        ("saatgut_los", "saatgut_los", False, "str"),
        ("saatgut_lieferant", "saatgut_lieferant", False, "str"),
        ("substrat", "substrat", False, "str"),
        ("substrat_menge", "substrat_menge", False, "decimal"),
        ("geplantes_ernte_datum", "geplantes_ernte_datum", False, "date"),
        ("ausschuss_menge_gramm", "ausschuss_menge_gramm", False, "decimal"),
        ("ausschuss_grund", "ausschuss_grund", False, "str"),
        ("notiz", "notiz", False, "str"),
    ],
}


def _coerce(value: Any, type_hint: str) -> Any:
    if value is None or value == "":
        return None
    try:
        if type_hint == "str":
            return str(value).strip() or None
        if type_hint == "int":
            return int(float(value))
        if type_hint == "decimal":
            return Decimal(str(value).replace(",", "."))
        if type_hint == "bool":
            s = str(value).strip().lower()
            return s in ("true", "1", "ja", "yes", "x")
        if type_hint == "date":
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            s = str(value).strip()
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None
        if type_hint.startswith("enum:"):
            valid = type_hint.split(":", 1)[1].split("|")
            v = str(value).strip().upper()
            return v if v in valid else None
    except (InvalidOperation, ValueError):
        return None
    return value


# ---- Beispieldaten je Entity ----------------------------------------------
#
# Landen auf einem eigenen Blatt "Beispiel", nicht im Datenblatt: eine
# ausgefüllte Zeile im Datenblatt würde beim ersten Upload als echter
# Datensatz importiert. Die Werte je Liste entsprechen der Spaltenreihenfolge
# in COLUMNS; Datumsangaben werden relativ zu heute erzeugt, damit das
# Beispiel nicht veraltet.


def _examples(entity: str) -> list[list[Any]]:
    from datetime import timedelta

    heute = date.today()
    d = lambda tage: (heute - timedelta(days=tage)).strftime("%d.%m.%Y")  # noqa: E731

    if entity == "customers":
        return [
            ["Gasthaus Sonne", "GASTRO", "bestellung@gasthaus-sonne.de", "089 1234567",
             "Hauptstraße 1, 80331 München", "DE123456789", "Liefertage Di + Fr"],
            ["BioMarkt Isartal", "HANDEL", "einkauf@biomarkt-isartal.de", "089 7654321",
             "Marktplatz 8, 82031 Grünwald", "", ""],
        ]
    if entity == "suppliers":
        return [
            ["Saatgut Müller GmbH", "kontakt@saatgut-mueller.de", "0711 998877",
             "Industriestraße 5, 70565 Stuttgart", "DE987654321", "SAATGUT", "ja",
             "DE-ÖKO-006", "Bio-Saatgut, Lieferzeit 5 Tage"],
        ]
    if entity == "seeds":
        return [
            # name, sorte, lieferant, keim, wachstum, ernte min/opt/max,
            # ertrag, verlust, saatgut/Einheit, cooling, cooling shelf, prozess
            ["Rucola", "Coltivata", "Saatgut Müller GmbH", 2, 5, 7, 8, 10, 250, 5, 30, 0, 0, "STANDARD"],
            ["Erbse", "Grünschnitt", "Saatgut Müller GmbH", 2, 7, 9, 10, 12, 700, 8, 250, 0, 0, "PLATTE"],
        ]
    if entity == "products":
        return [
            ["MG-RUC-100", "Rucola 100 g Schale", "Coltivata", "MICROGREEN", "4260123456789",
             "A-1001", "BIO", "Frische Rucola-Microgreens in der 100-g-Schale", "3.90",
             "REDUZIERT", 7, "nein", ""],
            ["MG-ERB-200", "Erbsengrün 200 g Schale", "Grünschnitt", "MICROGREEN", "",
             "A-1002", "BIO", "", "4.50", "REDUZIERT", 7, "nein", ""],
            # Pfandgebinde: 19 % statt 7 %, Pfandwert je Stück
            ["PFAND-KISTE", "Pfandkiste E2", "", "PACKAGING", "", "", "", "", "4.50",
             "STANDARD", "", "ja", "4.50"],
        ]
    if entity == "locations":
        return [
            ["KR-01", "Kühlraum 1", "KUEHLRAUM", "Fertigware nach der Ernte", 2, 6],
            ["LG-01", "Trockenlager", "LAGER", "Saatgut und Substrat", 15, 22],
        ]
    if entity == "order_history":
        return [
            # Zwei Zeilen mit gleicher externer Nummer = EINE Bestellung mit 2 Positionen
            ["EXT-1001", "Gasthaus Sonne", d(21), d(20), "MG-RUC-100", 12, "STK", "3.90", "GELIEFERT", ""],
            ["EXT-1001", "Gasthaus Sonne", d(21), d(20), "MG-ERB-200", 5, "STK", "4.50", "GELIEFERT", ""],
            ["EXT-1002", "BioMarkt Isartal", d(14), d(13), "MG-RUC-100", 30, "STK", "3.50", "GELIEFERT", ""],
        ]
    if entity == "grow_batches":
        return [
            ["Rucola", d(20), 12, "GEERNTET", "RU-2026-07", "R1-A", d(11), 40, ""],
            ["Erbse", d(3), 8, "", "", "R2-B", "", "", ""],
        ]
    return []


HINWEISE = {
    "order_history": (
        "Mehrere Zeilen mit derselben 'bestell_nr_extern' ergeben EINE Bestellung "
        "mit mehreren Positionen. Kunde und produkt_sku müssen bereits angelegt sein."
    ),
    "grow_batches": "Die Sorte muss vorher als Saatgut angelegt sein.",
    "products": (
        "Preise mit Punkt oder Komma — beides wird erkannt. Pfandgebinde: "
        "'pfand' = ja, 'pfandwert' je Stück, tax_rate = STANDARD (19 %)."
    ),
}


def _build_template(entity: str) -> bytes:
    cols = COLUMNS[entity]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="166534")

    def _write_header(sheet) -> None:
        for idx, (header, _attr, required, _type_hint) in enumerate(cols, start=1):
            cell = sheet.cell(row=1, column=idx, value=f"{header}{' *' if required else ''}")
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[cell.column_letter].width = max(15, len(header) + 4)

    wb = Workbook()
    # Blattname bewusst fix: der Import liest gezielt "Daten" und kann so nie
    # versehentlich das Beispielblatt einlesen.
    ws = wb.active
    ws.title = DATA_SHEET
    _write_header(ws)
    for idx, (_header, _attr, _required, type_hint) in enumerate(cols, start=1):
        ws.cell(row=2, column=idx, value=f"[{type_hint}]")

    beispiel = wb.create_sheet(EXAMPLE_SHEET)
    _write_header(beispiel)
    rows = _examples(entity)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            beispiel.cell(row=r_idx, column=c_idx, value=value)
    note_row = len(rows) + 3
    beispiel.cell(
        row=note_row,
        column=1,
        value=f"Nur Beispiel — bitte im Blatt '{DATA_SHEET}' erfassen. Spalten mit * sind Pflicht.",
    ).font = Font(italic=True)
    if entity in HINWEISE:
        beispiel.cell(row=note_row + 1, column=1, value=HINWEISE[entity]).font = Font(italic=True)

    wb.active = 0  # beim Öffnen steht der Cursor im Datenblatt
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template/{entity}")
def download_template(entity: str):
    if entity not in COLUMNS:
        raise HTTPException(status_code=404, detail=f"Unbekannte Entität: {entity}")
    data = _build_template(entity)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="template_{entity}.xlsx"'},
    )


def _parse_rows(file: UploadFile, entity: str) -> tuple[list[dict], list[str]]:
    """Liest XLSX, gibt (rows, errors) zurück."""
    cols = COLUMNS[entity]
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Datei konnte nicht gelesen werden: {e}")

    # Gezielt das Datenblatt lesen: Excel merkt sich das zuletzt angesehene
    # Blatt als aktives — sonst würde ein Blick ins Beispielblatt vor dem
    # Speichern die Beispieldaten importieren. Ältere Templates (ein Blatt)
    # fallen auf das aktive Blatt zurück.
    ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.active

    header_to_idx: dict[str, int] = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if cell:
            header_to_idx[str(cell).split(" ")[0].strip().lower()] = idx

    rows: list[dict] = []
    errors: list[str] = []
    for row_num, raw_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip type-hint row + completely empty rows
        if all(c is None or str(c).startswith("[") for c in raw_row):
            continue
        if all(c is None or str(c).strip() == "" for c in raw_row):
            continue
        record: dict[str, Any] = {}
        row_error = None
        for header, attr, required, type_hint in cols:
            col_idx = header_to_idx.get(header.lower())
            raw = raw_row[col_idx] if col_idx is not None and col_idx < len(raw_row) else None
            value = _coerce(raw, type_hint)
            if required and value is None:
                row_error = f"Zeile {row_num}: '{header}' fehlt"
                break
            record[attr] = value
        if row_error:
            errors.append(row_error)
            continue
        rows.append(record)
    return rows, errors


def _import_customers(db, rows: list[dict]) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        typ = CustomerType(r["typ"]) if r.get("typ") else CustomerType.GASTRO
        existing = db.execute(select(Customer).where(Customer.name == r["name"])).scalar_one_or_none()
        if existing:
            for k, v in r.items():
                if v is not None and k != "typ":
                    setattr(existing, k, v)
            existing.typ = typ
            updated += 1
        else:
            # Kundennummer wie in der Maske vergeben — ohne sie fehlt die
            # Zeile "Kundennummer" später auf Lieferschein und Rechnung.
            kunde = Customer(**{**r, "typ": typ})
            if not kunde.customer_number:
                kunde.customer_number = next_customer_number(db)
            db.add(kunde)
            db.flush()  # damit die nächste Zeile die vergebene Nummer sieht
            created += 1
    db.commit()
    return created, updated


def _import_suppliers(db, rows: list[dict]) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        existing = db.execute(select(Supplier).where(Supplier.name == r["name"])).scalar_one_or_none()
        if existing:
            for k, v in r.items():
                if v is not None:
                    setattr(existing, k, v)
            updated += 1
        else:
            db.add(Supplier(**r))
            created += 1
    db.commit()
    return created, updated


def _import_seeds(db, rows: list[dict]) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        existing = db.execute(select(Seed).where(Seed.name == r["name"])).scalar_one_or_none()
        if existing:
            for k, v in r.items():
                if v is not None:
                    setattr(existing, k, v)
            updated += 1
        else:
            db.add(Seed(**r))
            created += 1
    db.commit()
    return created, updated


def _import_products(db, rows: list[dict]) -> tuple[int, int]:
    default_unit = db.execute(select(UnitOfMeasure).where(UnitOfMeasure.code == "G")).scalar_one_or_none()
    if not default_unit:
        raise HTTPException(status_code=500, detail="Basiseinheit 'G' fehlt — bitte Stammdaten initialisieren")
    created = updated = 0
    for r in rows:
        category = ProductCategory(r["category"]) if r.get("category") else ProductCategory.MICROGREEN
        tax_rate = TaxRate(r["tax_rate"]) if r.get("tax_rate") else TaxRate.REDUZIERT
        existing = db.execute(select(Product).where(Product.sku == r["sku"])).scalar_one_or_none()
        payload = {**r, "category": category, "tax_rate": tax_rate}
        if existing:
            for k, v in payload.items():
                if v is not None:
                    setattr(existing, k, v)
            updated += 1
        else:
            db.add(Product(**{**payload, "base_unit_id": default_unit.id}))
            created += 1
    db.commit()
    return created, updated


def _import_locations(db, rows: list[dict]) -> tuple[int, int]:
    created = updated = 0
    for r in rows:
        loc_type = LocationType(r["location_type"]) if r.get("location_type") else LocationType.LAGER
        existing = db.execute(select(InventoryLocation).where(InventoryLocation.code == r["code"])).scalar_one_or_none()
        payload = {**r, "location_type": loc_type}
        if existing:
            for k, v in payload.items():
                if v is not None:
                    setattr(existing, k, v)
            updated += 1
        else:
            db.add(InventoryLocation(**payload))
            created += 1
    db.commit()
    return created, updated


def _generate_historic_order_number(db, order_date: date, used_numbers: set[str]) -> str:
    """Generiert BE-YYYYMMDD-NNNN für ein historisches Datum.

    Mit SELECT … FOR UPDATE auf den Row-Lock, damit zwei parallele
    Import-Läufe nicht beide dieselbe Nummer vergeben (sonst Unique-Verletzung
    + Rollback der einen Transaktion → Datenverlust).

    Berücksichtigt sowohl bereits existierende DB-Nummern als auch
    Nummern, die in dieser Import-Transaktion bereits vergeben wurden
    (used_numbers), damit Massenimporte kollisionsfrei bleiben."""
    prefix = f"BE-{order_date.strftime('%Y%m%d')}"
    last = db.execute(
        select(Order)
        .where(Order.order_number.like(f"{prefix}-%"))
        .order_by(Order.order_number.desc())
        .with_for_update()
        .limit(1)
    ).scalar_one_or_none()
    next_num = (int(last.order_number.split("-")[-1]) + 1) if last else 1
    while f"{prefix}-{next_num:04d}" in used_numbers:
        next_num += 1
    number = f"{prefix}-{next_num:04d}"
    used_numbers.add(number)
    return number


def _import_order_history(db, rows: list[dict]) -> tuple[int, int]:
    """Importiert historische Bestellungen für Forecast-Training.

    Gruppiert Zeilen nach `bestell_nr_extern` → eine Bestellung pro Gruppe.
    Idempotent über customer_reference (re-runs überspringen vorhandene)."""
    if not rows:
        return 0, 0

    # 1) Gruppieren
    groups: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        groups[r["bestell_nr_extern"]].append(r)

    # 2) Customer-Cache (case-insensitive + unicode-normalisierter Name-Lookup)
    # SQLite's func.lower() macht kein Unicode-Casefolding ("Ö" bleibt "Ö"),
    # daher laden wir einmal alle Kunden und vergleichen Python-seitig.
    import unicodedata

    def _normalize_name(s: str) -> str:
        return unicodedata.normalize("NFC", s.strip()).casefold()

    customers_by_name: dict[str, Customer] = {}
    products_by_sku: dict[str, Product] = {}

    # Alle Kunden einmal laden, normalisiert indexieren
    _all_customers = db.execute(select(Customer)).scalars().all()
    customer_index = {_normalize_name(c.name): c for c in _all_customers}

    def _get_customer(name: str) -> Optional[Customer]:
        key = _normalize_name(name)
        if key in customers_by_name:
            return customers_by_name[key]
        c = customer_index.get(key)
        if c:
            customers_by_name[key] = c
        return c

    def _get_product(sku: str) -> Optional[Product]:
        if sku in products_by_sku:
            return products_by_sku[sku]
        p = db.execute(select(Product).where(Product.sku == sku)).scalar_one_or_none()
        if p:
            products_by_sku[sku] = p
        return p

    created = skipped = 0
    used_numbers: set[str] = set()
    for ext_nr, group_rows in groups.items():
        # Idempotenz: gleicher customer_reference schon importiert → skip
        existing = db.execute(
            select(Order).where(Order.customer_reference == ext_nr).limit(1)
        ).scalar_one_or_none()
        if existing:
            skipped += 1
            continue

        head = group_rows[0]
        customer = _get_customer(head["kunde"])
        if not customer:
            raise HTTPException(
                status_code=400,
                detail=f"Bestellung '{ext_nr}': Kunde '{head['kunde']}' nicht gefunden — bitte zuerst Stammdaten importieren",
            )

        status_str = head.get("status") or "GELIEFERT"
        try:
            order_status = OrderStatus(status_str)
        except ValueError:
            order_status = OrderStatus.GELIEFERT

        order_number = _generate_historic_order_number(db, head["bestelldatum"], used_numbers)

        order = Order(
            order_number=order_number,
            customer_id=customer.id,
            customer_reference=ext_nr,
            order_date=datetime.combine(head["bestelldatum"], datetime.min.time()),
            requested_delivery_date=head["lieferdatum"],
            actual_delivery_date=head["lieferdatum"] if order_status == OrderStatus.GELIEFERT else None,
            status=order_status,
            currency="EUR",
            total_net=Decimal("0"),
            total_vat=Decimal("0"),
            total_gross=Decimal("0"),
            discount_percent=Decimal("0"),
            discount_amount=Decimal("0"),
        )
        db.add(order)
        db.flush()  # order.id verfügbar machen

        import json

        for position, line_row in enumerate(group_rows, start=1):
            product = _get_product(line_row["produkt_sku"])
            if not product:
                raise HTTPException(
                    status_code=400,
                    detail=f"Bestellung '{ext_nr}': SKU '{line_row['produkt_sku']}' nicht gefunden",
                )

            # Variable-Bundle-Sorten aus JSON-String (optional) auflösen
            vb_selections = None
            raw_sels = line_row.get("bundle_selections")
            if raw_sels:
                try:
                    parsed = json.loads(raw_sels) if isinstance(raw_sels, str) else raw_sels
                    vb_selections = []
                    for sel in parsed:
                        sku = sel.get("sku") or sel.get("product_sku")
                        qty = int(sel.get("quantity", 1) or 1)
                        if not sku:
                            continue
                        sort_prod = _get_product(sku)
                        if not sort_prod:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Bestellung '{ext_nr}': Bundle-Sorte '{sku}' nicht in Produkten gefunden",
                            )
                        vb_selections.append({"product_id": str(sort_prod.id), "quantity": qty})
                    if not vb_selections:
                        vb_selections = None
                except json.JSONDecodeError as e:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Bestellung '{ext_nr}': bundle_selections ist kein gültiges JSON: {e}",
                    )

            line = OrderLine(
                order_id=order.id,
                position=position,
                product_id=product.id,
                product_sku=product.sku,
                beschreibung=product.name,  # Snapshot des Produktnamens
                quantity=line_row["menge"],
                unit=line_row.get("einheit") or "g",
                unit_price=line_row["einzelpreis"],
                discount_percent=Decimal("0"),
                tax_rate=product.tax_rate or TaxRate.REDUZIERT,
                variable_bundle_selections=vb_selections,
            )
            line.calculate_line_totals()
            db.add(line)
            order.lines.append(line)

        order.calculate_totals()
        created += 1

    db.commit()
    return created, skipped


def _import_grow_batches(db, rows: list[dict]) -> tuple[int, int]:
    """Importiert Wachstumschargen (Go-Live: laufende + historische Aussaaten).

    - `sorte` wird gegen die Saatgut-Stammdaten aufgelöst (Name, case-insensitiv)
    - fehlende Saatgut-Charge wird als 'IMPORT-...'-Marker angelegt (Menge 0)
    - Erntefenster wird aus den Sortenparametern berechnet (Tage ab Aussaat)
    - Status-Default: GEERNTET wenn ernte_datum gesetzt, sonst nach Alter
    - Idempotent über (Sorte, Aussaatdatum, Kistenzahl): vorhandene übersprungen
    """
    import unicodedata
    from datetime import timedelta, date as _date
    from app.models.seed import Seed, SeedBatch
    from app.models.production import GrowBatch, GrowBatchStatus, Harvest

    if not rows:
        return 0, 0

    def _norm(s: str) -> str:
        return unicodedata.normalize("NFC", s.strip()).casefold()

    seeds_index = {_norm(s.name): s for s in db.execute(select(Seed)).scalars().all()}
    created = skipped = 0

    for r in rows:
        seed = seeds_index.get(_norm(r["sorte"]))
        if seed is None:
            raise ValueError(f"Sorte '{r['sorte']}' nicht gefunden — bitte zuerst als Saatgut anlegen")

        aussaat = r["aussaat_datum"]
        trays = r["tray_anzahl"]

        # Idempotenz: gleiche Sorte + Aussaat + Kistenzahl gilt als bereits importiert
        existing = db.execute(
            select(GrowBatch)
            .join(SeedBatch, GrowBatch.seed_batch_id == SeedBatch.id)
            .where(
                SeedBatch.seed_id == seed.id,
                GrowBatch.aussaat_datum == aussaat,
                GrowBatch.tray_anzahl == trays,
            )
        ).scalars().first()
        if existing:
            skipped += 1
            continue

        # Saatgut-Charge: vorhandene per Nummer, sonst Import-Marker
        charge_nummer = r.get("charge_nummer") or f"IMPORT-{seed.name[:20]}"
        seed_batch = db.execute(
            select(SeedBatch).where(SeedBatch.seed_id == seed.id, SeedBatch.charge_nummer == charge_nummer)
        ).scalar_one_or_none()
        if seed_batch is None:
            seed_batch = SeedBatch(
                seed_id=seed.id, charge_nummer=charge_nummer,
                menge_gramm=Decimal("0"), verbleibend_gramm=Decimal("0"),
            )
            db.add(seed_batch)
            db.flush()

        status = r.get("status")
        if not status:
            if r.get("ernte_datum"):
                status = "GEERNTET"
            elif aussaat + timedelta(days=seed.erntefenster_max_tage) < _date.today():
                status = "GEERNTET"
            elif aussaat + timedelta(days=seed.erntefenster_min_tage) <= _date.today():
                status = "ERNTEREIF"
            else:
                status = "WACHSTUM"

        batch = GrowBatch(
            seed_batch_id=seed_batch.id,
            tray_anzahl=trays,
            aussaat_datum=aussaat,
            erwartete_ernte_min=aussaat + timedelta(days=seed.erntefenster_min_tage),
            erwartete_ernte_optimal=aussaat + timedelta(days=seed.erntefenster_optimal_tage),
            erwartete_ernte_max=aussaat + timedelta(days=seed.erntefenster_max_tage),
            status=GrowBatchStatus(status),
            regal_position=r.get("regal_position"),
            notizen="Go-Live-Import",
        )
        db.add(batch)
        db.flush()

        # Optionale Ernte-Daten
        if r.get("ernte_datum") and (r.get("ernte_menge_stueck") or r.get("ernte_menge_gramm")):
            stueck = r.get("ernte_menge_stueck")
            db.add(Harvest(
                grow_batch_id=batch.id,
                ernte_datum=r["ernte_datum"],
                einheit="STK" if stueck else "G",
                menge_stueck=stueck,
                menge_gramm=Decimal("0") if stueck else r.get("ernte_menge_gramm"),
            ))

        created += 1

    db.commit()
    return created, skipped


IMPORTERS = {
    "customers": _import_customers,
    "suppliers": _import_suppliers,
    "seeds": _import_seeds,
    "products": _import_products,
    "locations": _import_locations,
    "order_history": _import_order_history,
    "grow_batches": _import_grow_batches,
}


@router.post("/{entity}")
async def import_entity(entity: str, db: DBSession, file: UploadFile = File(...)):
    if entity not in IMPORTERS:
        raise HTTPException(status_code=404, detail=f"Unbekannte Entität: {entity}")
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Nur .xlsx/.xlsm Dateien werden unterstützt")
    rows, parse_errors = _parse_rows(file, entity)
    if not rows and parse_errors:
        return {"created": 0, "updated": 0, "errors": parse_errors}
    try:
        created, updated = IMPORTERS[entity](db, rows)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Import fehlgeschlagen: {e}")
    return {"created": created, "updated": updated, "errors": parse_errors}


# =====================================================================
# Historien-Import v2 (Warenfluss-Release, AP3)
#
# Zweistufig: /grow-batches/validate liefert einen Zeilenreport, erst
# /grow-batches/commit schreibt — als Import-Lauf, der über /runs/{id}
# in einem Schritt rückrollbar ist, solange keine Folgebelege daran hängen.
#
# Festlegung Bestandswirkung: historische Bewegungen buchen ins JOURNAL
# (movement_date = historisches Datum), nicht auf den Ist-Bestand. Der
# Ist-Bestand ist gezählte Gegenwart; die Sorte hängt an der Bewegung über
# grow_batch → seed_batch → seed, ein Bestandseintrag ist nicht nötig.
# =====================================================================

from datetime import time as _time, timedelta as _timedelta, timezone as _timezone

from app.models.import_run import ImportRun
from app.models.inventory import (
    InventoryItemType, InventoryMovement, MovementType, PackagingInventory,
)
from app.models.production import GrowBatch, GrowBatchStatus, Harvest
from app.models.seed import SeedBatch


def _parse_rows_mit_zeilen(file: UploadFile, entity: str) -> list[tuple[int, Optional[dict], Optional[str]]]:
    """Wie _parse_rows, aber je Zeile (zeilennummer, record, fehler).

    Für den Validierungsreport: ein Fehler bricht nicht den ganzen Import ab,
    sondern markiert genau seine Zeile.
    """
    cols = COLUMNS[entity]
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Datei konnte nicht gelesen werden: {e}")

    ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.active
    header_to_idx: dict[str, int] = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if cell:
            header_to_idx[str(cell).split(" ")[0].strip().lower()] = idx

    ergebnis: list[tuple[int, Optional[dict], Optional[str]]] = []
    for row_num, raw_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(c is None or str(c).startswith("[") for c in raw_row):
            continue
        if all(c is None or str(c).strip() == "" for c in raw_row):
            continue
        record: dict[str, Any] = {}
        fehler: Optional[str] = None
        for header, attr, required, type_hint in cols:
            col_idx = header_to_idx.get(header.lower())
            raw = raw_row[col_idx] if col_idx is not None and col_idx < len(raw_row) else None
            value = _coerce(raw, type_hint)
            if required and value is None:
                fehler = f"'{header}' fehlt"
                break
            record[attr] = value
        ergebnis.append((row_num, None if fehler else record, fehler))
    return ergebnis


def _norm_name(s: str) -> str:
    import unicodedata
    return unicodedata.normalize("NFC", str(s).strip()).casefold()


def _chargen_report(db, geparste) -> dict:
    """Zeilenreport für den Chargen-Import (R3.3): OK / WARNUNG / FEHLER."""
    seeds_index = {_norm_name(s.name): s for s in db.execute(select(Seed)).scalars().all()}
    substrat_index = {
        _norm_name(p.name): p
        for p in db.execute(
            select(PackagingInventory).where(PackagingInventory.article_type == "SUBSTRAT")
        ).scalars().all()
    }
    vorhandene_extern = {
        nr for (nr,) in db.execute(
            select(GrowBatch.externe_chargennummer).where(GrowBatch.externe_chargennummer.isnot(None))
        ).all()
    }

    zeilen = []
    fehlende_sorten: list[str] = []
    fehlende_substrate: list[str] = []
    gesehen_extern: set[str] = set()

    for zeile, record, parse_fehler in geparste:
        if parse_fehler:
            zeilen.append({"zeile": zeile, "status": "FEHLER", "meldung": parse_fehler})
            continue

        meldungen: list[str] = []
        status = "OK"

        seed = seeds_index.get(_norm_name(record["sorte"]))
        if seed is None:
            status = "FEHLER"
            meldungen.append(f"Sorte '{record['sorte']}' nicht gefunden — bitte zuerst als Saatgut anlegen")
            if record["sorte"] not in fehlende_sorten:
                fehlende_sorten.append(record["sorte"])

        extern = record.get("externe_chargennummer")
        if extern:
            if extern in gesehen_extern:
                status = "FEHLER"
                meldungen.append(f"externe_chargennummer '{extern}' kommt in der Datei doppelt vor")
            gesehen_extern.add(extern)
            if extern in vorhandene_extern and status != "FEHLER":
                meldungen.append("bereits importiert — wird übersprungen")

        if record.get("ernte_menge_gramm") and not record.get("ernte_datum"):
            status = "FEHLER"
            meldungen.append("ernte_menge_gramm ohne ernte_datum")

        if status != "FEHLER":
            if seed is not None and not record.get("saatgut_gramm") and not getattr(seed, "saatgut_pro_einheit_gramm", None):
                status = "WARNUNG"
                meldungen.append("kein saatgut_gramm und Sorte ohne Saatgutmenge je Kiste — es wird kein Verbrauch gebucht")
            substrat = record.get("substrat")
            if substrat and _norm_name(substrat) not in substrat_index:
                status = "WARNUNG" if status == "OK" else status
                meldungen.append(f"Substrat '{substrat}' nicht im Lager angelegt — Verbrauch wird nicht gebucht")
                if substrat not in fehlende_substrate:
                    fehlende_substrate.append(substrat)

        zeilen.append({
            "zeile": zeile, "status": status,
            "meldung": "; ".join(meldungen) if meldungen else "in Ordnung",
            "sorte": record.get("sorte"),
        })

    return {
        "zeilen": zeilen,
        "zusammenfassung": {
            "ok": sum(1 for z in zeilen if z["status"] == "OK"),
            "warnung": sum(1 for z in zeilen if z["status"] == "WARNUNG"),
            "fehler": sum(1 for z in zeilen if z["status"] == "FEHLER"),
        },
        "fehlende_sorten": fehlende_sorten,
        "fehlende_substrate": fehlende_substrate,
    }


def _historische_bewegung(run_id, *, movement_type, item_type, quantity, unit,
                          movement_date, grow_batch_id=None, harvest_id=None,
                          packaging_id=None, reason=None) -> InventoryMovement:
    """Journalbuchung mit historischem Datum, ohne Ist-Bestand anzufassen.

    quantity_before/after bleiben 0: zum historischen Zeitpunkt ist kein
    Bestandsstand rekonstruierbar — die Auswertungen summieren quantity.
    """
    return InventoryMovement(
        movement_type=movement_type,
        item_type=item_type,
        quantity=quantity,
        unit=unit,
        quantity_before=Decimal("0"),
        quantity_after=Decimal("0"),
        movement_date=datetime.combine(movement_date, _time.min),
        grow_batch_id=grow_batch_id,
        harvest_id=harvest_id,
        packaging_id=packaging_id,
        reason=reason,
        reference_number=f"IMPORT:{run_id}",
    )


@router.post("/grow-batches/validate")
async def validate_grow_batch_import(db: DBSession, file: UploadFile = File(...)):
    """Dry-Run (R3.3): prüft die Datei Zeile für Zeile, schreibt nichts."""
    geparste = _parse_rows_mit_zeilen(file, "grow_batches")
    return _chargen_report(db, geparste)


@router.post("/grow-batches/commit", status_code=201)
async def commit_grow_batch_import(
    db: DBSession,
    file: UploadFile = File(...),
    lagerbewegungen: bool = True,
):
    """Führt den Chargen-Import aus (R3.4–R3.6).

    Enthält die Datei Fehlerzeilen, wird komplett abgelehnt — der Report
    steht in der Fehlermeldung. Warnungen importieren.
    """
    geparste = _parse_rows_mit_zeilen(file, "grow_batches")
    report = _chargen_report(db, geparste)
    if report["zusammenfassung"]["fehler"] > 0:
        raise HTTPException(status_code=400, detail={
            "meldung": "Datei enthält Fehlerzeilen — nichts wurde importiert.",
            "report": report,
        })

    seeds_index = {_norm_name(s.name): s for s in db.execute(select(Seed)).scalars().all()}
    substrat_index = {
        _norm_name(p.name): p
        for p in db.execute(
            select(PackagingInventory).where(PackagingInventory.article_type == "SUBSTRAT")
        ).scalars().all()
    }
    vorhandene_extern = {
        nr for (nr,) in db.execute(
            select(GrowBatch.externe_chargennummer).where(GrowBatch.externe_chargennummer.isnot(None))
        ).all()
    }

    run = ImportRun(entity="grow_batches", filename=file.filename)
    db.add(run)
    db.flush()

    created = skipped = bewegungen = 0

    for _zeile, record, _fehler in geparste:
        seed = seeds_index[_norm_name(record["sorte"])]
        aussaat = record["aussaat_datum"]
        trays = record["tray_anzahl"]
        extern = record.get("externe_chargennummer")

        # Idempotenz: primär externe Chargennummer, ersatzweise fachlicher Schlüssel
        if extern and extern in vorhandene_extern:
            skipped += 1
            continue
        if not extern:
            doppelt = db.execute(
                select(GrowBatch)
                .join(SeedBatch, GrowBatch.seed_batch_id == SeedBatch.id)
                .where(SeedBatch.seed_id == seed.id,
                       GrowBatch.aussaat_datum == aussaat,
                       GrowBatch.tray_anzahl == trays)
            ).scalars().first()
            if doppelt:
                skipped += 1
                continue

        # Saatgut-Charge: Los aus der Datei oder Import-Marker (Menge 0)
        charge_nummer = record.get("saatgut_los") or record.get("charge_nummer") or f"IMPORT-{seed.name[:20]}"
        seed_batch = db.execute(
            select(SeedBatch).where(SeedBatch.seed_id == seed.id,
                                    SeedBatch.charge_nummer == charge_nummer)
        ).scalar_one_or_none()
        if seed_batch is None:
            # Lieferant steht als Freitext im Lieferschein-Feld — die Charge
            # hat kein eigenes Lieferantenfeld, und für den Auditnachweis
            # zählt, dass Los und Herkunft am Datensatz ablesbar sind.
            seed_batch = SeedBatch(
                seed_id=seed.id, charge_nummer=charge_nummer,
                menge_gramm=Decimal("0"), verbleibend_gramm=Decimal("0"),
                lieferschein_nr=record.get("saatgut_lieferant"),
            )
            db.add(seed_batch)
            db.flush()

        status = record.get("status")
        if not status:
            if record.get("ernte_datum"):
                status = "GEERNTET"
            elif aussaat + _timedelta(days=seed.erntefenster_max_tage) < date.today():
                status = "GEERNTET"
            elif aussaat + _timedelta(days=seed.erntefenster_min_tage) <= date.today():
                status = "ERNTEREIF"
            else:
                status = "WACHSTUM"

        batch = GrowBatch(
            seed_batch_id=seed_batch.id,
            tray_anzahl=trays,
            aussaat_datum=aussaat,
            erwartete_ernte_min=aussaat + _timedelta(days=seed.erntefenster_min_tage),
            erwartete_ernte_optimal=record.get("geplantes_ernte_datum")
                or aussaat + _timedelta(days=seed.erntefenster_optimal_tage),
            erwartete_ernte_max=aussaat + _timedelta(days=seed.erntefenster_max_tage),
            status=GrowBatchStatus(status),
            regal_position=record.get("regal_position"),
            notizen=record.get("notiz") or "Historien-Import",
            source="import",
            import_run_id=run.id,
            externe_chargennummer=extern,
        )
        db.add(batch)
        db.flush()

        harvest = None
        if record.get("ernte_datum") and (record.get("ernte_menge_stueck") or record.get("ernte_menge_gramm")):
            stueck = record.get("ernte_menge_stueck")
            harvest = Harvest(
                grow_batch_id=batch.id,
                ernte_datum=record["ernte_datum"],
                einheit="STK" if stueck else "G",
                menge_stueck=stueck,
                menge_gramm=Decimal("0") if stueck else record.get("ernte_menge_gramm"),
                import_run_id=run.id,
            )
            db.add(harvest)
            db.flush()

        if lagerbewegungen:
            saatgut = record.get("saatgut_gramm") or (
                (getattr(seed, "saatgut_pro_einheit_gramm", None) or 0) and
                Decimal(str(seed.saatgut_pro_einheit_gramm)) * trays
            )
            if saatgut:
                db.add(_historische_bewegung(
                    run.id, movement_type=MovementType.PRODUKTION,
                    item_type=InventoryItemType.SAATGUT,
                    quantity=-Decimal(str(saatgut)), unit="G",
                    movement_date=aussaat, grow_batch_id=batch.id,
                    reason="Aussaat (Historien-Import)",
                ))
                bewegungen += 1
            if harvest is not None and record.get("ernte_menge_gramm"):
                db.add(_historische_bewegung(
                    run.id, movement_type=MovementType.ERNTE,
                    item_type=InventoryItemType.FERTIGWARE,
                    quantity=Decimal(str(record["ernte_menge_gramm"])), unit="G",
                    movement_date=record["ernte_datum"],
                    grow_batch_id=batch.id, harvest_id=harvest.id,
                    reason="Ernte (Historien-Import)",
                ))
                bewegungen += 1
            if record.get("ausschuss_menge_gramm"):
                db.add(_historische_bewegung(
                    run.id, movement_type=MovementType.VERLUST,
                    item_type=InventoryItemType.FERTIGWARE,
                    quantity=-Decimal(str(record["ausschuss_menge_gramm"])), unit="G",
                    movement_date=record.get("ernte_datum") or aussaat,
                    grow_batch_id=batch.id,
                    reason=record.get("ausschuss_grund") or "Ausschuss (Historien-Import)",
                ))
                bewegungen += 1
            substrat = record.get("substrat")
            if substrat and record.get("substrat_menge"):
                artikel = substrat_index.get(_norm_name(substrat))
                if artikel is not None:
                    db.add(_historische_bewegung(
                        run.id, movement_type=MovementType.PRODUKTION,
                        item_type=InventoryItemType.SUBSTRAT,
                        quantity=-Decimal(str(record["substrat_menge"])),
                        unit=artikel.unit,
                        movement_date=aussaat, grow_batch_id=batch.id,
                        packaging_id=artikel.id,
                        reason="Substratverbrauch (Historien-Import)",
                    ))
                    bewegungen += 1

        if extern:
            vorhandene_extern.add(extern)
        created += 1

    run.rows_created = created
    run.rows_skipped = skipped
    run.movements_created = bewegungen
    db.commit()

    return {
        "import_run_id": str(run.id),
        "created": created,
        "skipped": skipped,
        "movements": bewegungen,
        "report": report,
    }


@router.get("/runs")
def list_import_runs(db: DBSession):
    """Alle Import-Läufe, neueste zuerst."""
    runs = db.execute(select(ImportRun).order_by(ImportRun.created_at.desc())).scalars().all()
    return [{
        "id": str(r.id), "entity": r.entity, "filename": r.filename,
        "status": r.status, "created": r.rows_created, "skipped": r.rows_skipped,
        "movements": r.movements_created, "created_at": r.created_at.isoformat(),
    } for r in runs]


@router.delete("/runs/{run_id}")
def rollback_import_run(run_id: UUID, db: DBSession):
    """Rollt einen Import-Lauf komplett zurück (R3.6).

    Blockiert, sobald Folgebelege an importierten Chargen hängen — eine
    später erfasste echte Ernte oder Buchung darf nicht mit verschwinden.
    Import-Marker-Saatgutchargen (Menge 0) bleiben stehen; sie sind
    wirkungslos und können weitere Läufe tragen.
    """
    run = db.get(ImportRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Import-Lauf nicht gefunden")
    if run.status == "ZURUECKGEROLLT":
        raise HTTPException(status_code=409, detail="Lauf wurde bereits zurückgerollt")

    batches = db.execute(
        select(GrowBatch).where(GrowBatch.import_run_id == run_id)
    ).scalars().all()
    batch_ids = [b.id for b in batches]

    if batch_ids:
        fremde_ernte = db.execute(
            select(Harvest).where(
                Harvest.grow_batch_id.in_(batch_ids),
                (Harvest.import_run_id.is_(None)) | (Harvest.import_run_id != run_id),
            )
        ).scalars().first()
        if fremde_ernte:
            raise HTTPException(
                status_code=409,
                detail="An importierten Chargen hängen inzwischen echte Ernten — "
                       "Rollback nicht möglich.",
            )
        fremde_bewegung = db.execute(
            select(InventoryMovement).where(
                InventoryMovement.grow_batch_id.in_(batch_ids),
                (InventoryMovement.reference_number.is_(None))
                | (InventoryMovement.reference_number != f"IMPORT:{run_id}"),
            )
        ).scalars().first()
        if fremde_bewegung:
            raise HTTPException(
                status_code=409,
                detail="An importierten Chargen hängen inzwischen echte Lagerbewegungen — "
                       "Rollback nicht möglich.",
            )

        for bewegung in db.execute(
            select(InventoryMovement)
            .where(InventoryMovement.reference_number == f"IMPORT:{run_id}")
        ).scalars().all():
            db.delete(bewegung)
        for ernte in db.execute(
            select(Harvest).where(Harvest.import_run_id == run_id)
        ).scalars().all():
            db.delete(ernte)
        for batch in batches:
            db.delete(batch)

    run.status = "ZURUECKGEROLLT"
    db.commit()
    return {"status": "ZURUECKGEROLLT", "geloescht": len(batch_ids)}
