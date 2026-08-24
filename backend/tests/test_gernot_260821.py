"""
Tests: Gernot-Feedback vom 21.08.2026.

Abgedeckt:
- Excel-Templates brauchen eine ausgefüllte Beispielzeile ("damit ich genau
  weiß was ich einfüllen muss") — ohne dass das Beispiel beim Re-Upload als
  echte Datenzeile importiert wird.
- Punkt 1: Abo deaktivieren/bearbeiten schlug bei Produkt-Abos fehl.
"""
import io
from datetime import date, timedelta
from decimal import Decimal
from uuid import UUID

import pytest
from openpyxl import load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def pdf_text(raw: bytes) -> str:
    """Extrahiert Text aus einem ReportLab-PDF (ASCII85 + Flate)."""
    from tests.test_documents_preise import _pdf_text

    return _pdf_text(raw).decode("latin-1", errors="ignore")

ENTITIES = ["customers", "suppliers", "seeds", "products", "locations", "order_history", "grow_batches"]


def _upload(client, entity: str, content: bytes):
    return client.post(
        f"/api/v1/imports/{entity}",
        files={"file": (f"template_{entity}.xlsx", content, XLSX_MIME)},
    )


class TestImportTemplates:
    @pytest.mark.parametrize("entity", ENTITIES)
    def test_template_hat_ausgefuelltes_beispielblatt(self, client, entity):
        r = client.get(f"/api/v1/imports/template/{entity}")
        assert r.status_code == 200, r.text

        wb = load_workbook(io.BytesIO(r.content))
        assert "Beispiel" in wb.sheetnames, f"{entity}: Beispielblatt fehlt"

        ws = wb["Beispiel"]
        header = [c.value for c in ws[1]]
        beispiel = [c.value for c in ws[2]]

        # Jede Pflichtspalte (Header endet auf "*") muss im Beispiel gefüllt sein
        for idx, h in enumerate(header):
            if h and str(h).rstrip().endswith("*"):
                assert beispiel[idx] not in (None, ""), f"{entity}: Pflichtfeld '{h}' im Beispiel leer"

    @pytest.mark.parametrize("entity", ENTITIES)
    def test_datenblatt_ist_leer_und_heisst_daten(self, client, entity):
        r = client.get(f"/api/v1/imports/template/{entity}")
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames[0] == "Daten", f"{entity}: erstes Blatt muss 'Daten' heißen"
        assert wb.active.title == "Daten", f"{entity}: 'Daten' muss beim Öffnen aktiv sein"

    @pytest.mark.parametrize("entity", ["customers", "order_history", "grow_batches"])
    def test_unveraendertes_template_importiert_nichts(self, client, entity):
        """Das Beispiel darf nicht versehentlich als Datensatz landen."""
        tpl = client.get(f"/api/v1/imports/template/{entity}").content
        r = _upload(client, entity, tpl)
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 0, f"{entity}: Beispielzeile wurde importiert"
        assert r.json()["errors"] == []

    def test_beispielblatt_wird_nie_als_datenblatt_gelesen(self, client):
        """Excel speichert das zuletzt angesehene Blatt als aktiv.

        Wer im Template auf 'Beispiel' klickt und speichert, darf damit nicht
        die Beispieldaten importieren."""
        tpl = client.get("/api/v1/imports/template/customers").content
        wb = load_workbook(io.BytesIO(tpl))
        wb.active = wb.sheetnames.index("Beispiel")
        buf = io.BytesIO()
        wb.save(buf)

        r = _upload(client, "customers", buf.getvalue())
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 0, "Beispielblatt wurde als Datenblatt importiert"

    def test_altes_einblatt_template_bleibt_importierbar(self, client):
        """Vor dem Beispielblatt ausgegebene Templates müssen weiter funktionieren."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "customers"
        ws.append(["name *", "typ *", "email"])
        ws.append(["[str]", "[enum:GASTRO|HANDEL|GEWERBE|PRIVAT]", "[str]"])
        ws.append(["Altes Template GmbH", "GASTRO", "alt@example.com"])
        buf = io.BytesIO()
        wb.save(buf)

        r = _upload(client, "customers", buf.getvalue())
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 1


@pytest.fixture()
def base_unit(client):
    """Basiseinheit 'G' — Produktanlage verlangt sie, es gibt keinen API-Weg dorthin."""
    from app.models.unit import UnitOfMeasure, UnitCategory
    from tests.conftest import TestingSessionLocal

    db = TestingSessionLocal()
    try:
        if not db.query(UnitOfMeasure).filter_by(code="G").first():
            db.add(UnitOfMeasure(code="G", name="Gramm", symbol="g",
                                 category=UnitCategory.WEIGHT, is_base_unit=True))
            db.commit()
    finally:
        db.close()


class TestAbonnementBearbeiten:
    """Punkt 1: 'Fehler beim Deaktivieren' / 'Fehler beim Aktualisieren'.

    Beide Buttons in der UI rufen denselben PATCH-Endpunkt auf. Gernots Abo
    hängt an einem Produkt ('BIO Snackbox | Amaranth') und hat kein Saatgut.
    """

    def _kunde(self, client):
        return client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()

    def _produkt_abo(self, client):
        kunde = self._kunde(client)
        produkt = client.post("/api/v1/products", json={
            "sku": "SNACK-AMA", "name": "BIO Snackbox | Amaranth",
            "category": "BUNDLE", "base_price": 4.50,
        }).json()
        r = client.post("/api/v1/sales/subscriptions", json={
            "kunde_id": kunde["id"],
            "product_id": produkt["id"],
            "menge": 2,
            "einheit": "STUECK",
            "intervall": "WOECHENTLICH",
            "liefertage": [0],
            "gueltig_von": date.today().isoformat(),
        })
        assert r.status_code == 201, r.text
        return kunde, produkt, r.json()

    def test_produkt_abo_deaktivieren(self, client, base_unit):
        _, _, abo = self._produkt_abo(client)

        r = client.patch(f"/api/v1/sales/subscriptions/{abo['id']}", json={"aktiv": False})

        assert r.status_code == 200, r.text
        assert r.json()["aktiv"] is False

    def test_produkt_abo_bearbeiten_behaelt_produktnamen(self, client, base_unit):
        kunde, produkt, abo = self._produkt_abo(client)

        r = client.patch(f"/api/v1/sales/subscriptions/{abo['id']}", json={"menge": 5})

        assert r.status_code == 200, r.text
        body = r.json()
        assert float(body["menge"]) == 5
        # Ohne Namen zeigt die UI die nackte UUID an
        assert body["product_name"] == produkt["name"]
        assert body["kunde_name"] == kunde["name"]

    def test_saatgut_abo_bleibt_bearbeitbar(self, client):
        kunde = self._kunde(client)
        saatgut = client.post("/api/v1/seeds", json={
            "name": "Gartenkresse",
            "keimdauer_tage": 3,
            "wachstumsdauer_tage": 3,
            "erntefenster_min_tage": 6,
            "erntefenster_optimal_tage": 7,
            "erntefenster_max_tage": 8,
            "ertrag_gramm_pro_tray": 350,
            "verlustquote_prozent": 5.0,
        }).json()
        r = client.post("/api/v1/sales/subscriptions", json={
            "kunde_id": kunde["id"],
            "seed_id": saatgut["id"],
            "menge": 100,
            "einheit": "G",
            "intervall": "TAEGLICH",
            "gueltig_von": date.today().isoformat(),
        })
        assert r.status_code == 201, r.text

        r = client.patch(f"/api/v1/sales/subscriptions/{r.json()['id']}", json={"aktiv": False})

        assert r.status_code == 200, r.text
        assert r.json()["seed_name"] == saatgut["name"]


class TestUeberfaelligeRechnungen:
    """Punkt 2c: Rechnungen sind unter 'Alle' als überfällig markiert, fehlen
    aber im Reiter 'Überfällig' und in der Kennzahl oben."""

    def _ueberfaellige_rechnung(self, client):
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()
        r = client.post("/api/v1/invoices", json={
            "customer_id": kunde["id"],
            "invoice_date": (date.today() - timedelta(days=60)).isoformat(),
            "due_date": (date.today() - timedelta(days=30)).isoformat(),
        })
        assert r.status_code == 201, r.text
        rechnung = r.json()

        from app.models.invoice import Invoice
        from app.models.enums import InvoiceStatus
        from tests.conftest import TestingSessionLocal
        db = TestingSessionLocal()
        try:
            inv = db.get(Invoice, UUID(rechnung["id"]))
            inv.status = InvoiceStatus.OFFEN
            db.commit()
        finally:
            db.close()
        return rechnung

    def test_ueberfaellige_bleibt_beim_zweiten_aufruf_sichtbar(self, client):
        """Der erste Aufruf markiert — danach war die Liste jedes Mal leer."""
        rechnung = self._ueberfaellige_rechnung(client)

        erste = client.get("/api/v1/invoices/overdue")
        assert erste.status_code == 200, erste.text
        assert [i["id"] for i in erste.json()] == [rechnung["id"]]

        zweite = client.get("/api/v1/invoices/overdue")
        assert zweite.status_code == 200, zweite.text
        assert [i["id"] for i in zweite.json()] == [rechnung["id"]], \
            "Nach dem Markieren verschwindet die Rechnung aus Reiter und Kennzahl"

    def test_bezahlte_rechnung_ist_nicht_ueberfaellig(self, client):
        rechnung = self._ueberfaellige_rechnung(client)
        client.get("/api/v1/invoices/overdue")

        from app.models.invoice import Invoice
        from app.models.enums import InvoiceStatus
        from tests.conftest import TestingSessionLocal
        db = TestingSessionLocal()
        try:
            inv = db.get(Invoice, UUID(rechnung["id"]))
            inv.status = InvoiceStatus.BEZAHLT
            db.commit()
        finally:
            db.close()

        assert client.get("/api/v1/invoices/overdue").json() == []


class TestBestandskorrektur:
    """Punkt 5d: 'Bestandskorrektur klappt nicht — Fehler bei der Korrektur'."""

    @pytest.fixture()
    def saatgut_charge(self, client):
        saatgut = client.post("/api/v1/seeds", json={
            "name": "Rucola",
            "keimdauer_tage": 3,
            "wachstumsdauer_tage": 5,
            "erntefenster_min_tage": 7,
            "erntefenster_optimal_tage": 8,
            "erntefenster_max_tage": 10,
            "ertrag_gramm_pro_tray": 300,
            "verlustquote_prozent": 5.0,
        }).json()

        from app.models.inventory import SeedInventory
        from tests.conftest import TestingSessionLocal
        db = TestingSessionLocal()
        try:
            charge = SeedInventory(
                seed_id=UUID(saatgut["id"]),
                batch_number="CH-2026-001",
                initial_quantity_kg=Decimal("5.000"),
                current_quantity_kg=Decimal("5.000"),
                received_date=date.today(),
            )
            db.add(charge)
            db.commit()
            return str(charge.id)
        finally:
            db.close()

    def test_korrektur_bucht_neuen_bestand(self, client, saatgut_charge):
        r = client.post("/api/v1/inventory/correction", params={
            "inventory_id": saatgut_charge,
            "inventory_type": "SAATGUT",
            "actual_quantity": 4.2,
            "reason": "Zählfehler",
        })

        assert r.status_code == 200, r.text
        assert float(r.json()["new_quantity"]) == pytest.approx(4.2)

    def test_korrektur_schreibt_lagerbewegung(self, client, saatgut_charge):
        client.post("/api/v1/inventory/correction", params={
            "inventory_id": saatgut_charge,
            "inventory_type": "SAATGUT",
            "actual_quantity": 4.2,
            "reason": "Zählfehler",
        })

        from app.models.inventory import InventoryMovement
        from tests.conftest import TestingSessionLocal
        db = TestingSessionLocal()
        try:
            bewegungen = db.query(InventoryMovement).all()
            assert len(bewegungen) == 1
            assert bewegungen[0].seed_inventory_id == UUID(saatgut_charge)
            assert bewegungen[0].reason == "Korrektur: Zählfehler"
            assert bewegungen[0].unit, "Einheit ist Pflichtfeld in der Bewegung"
        finally:
            db.close()

    def test_korrektur_ohne_aenderung_ist_kein_fehler(self, client, saatgut_charge):
        r = client.post("/api/v1/inventory/correction", params={
            "inventory_id": saatgut_charge,
            "inventory_type": "SAATGUT",
            "actual_quantity": 5.0,
            "reason": "Nachgezählt",
        })
        assert r.status_code == 200, r.text


class TestBelegdaten:
    """Punkt 3c: Kundennummer und Adresszusatz fehlen auf dem Beleg."""

    def test_importierter_kunde_bekommt_kundennummer(self, client):
        """Über Excel angelegte Kunden blieben ohne Nummer — die Nummer steht
        dann auf keinem Beleg (Lieferschein lässt die Zeile ganz weg)."""
        tpl = client.get("/api/v1/imports/template/customers").content
        wb = load_workbook(io.BytesIO(tpl))
        ws = wb["Daten"]
        header = [str(c.value or "") for c in ws[1]]
        zeile = [None] * len(header)
        zeile[header.index("name *")] = "RATIONAL AG"
        zeile[header.index("typ *")] = "GEWERBE"
        ws.append(zeile)
        buf = io.BytesIO()
        wb.save(buf)

        r = _upload(client, "customers", buf.getvalue())
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 1

        kunden = client.get("/api/v1/sales/customers").json()["items"]
        importiert = next(k for k in kunden if k["name"] == "RATIONAL AG")
        assert importiert["customer_number"], "Kundennummer fehlt nach dem Import"
        assert importiert["customer_number"].startswith("KD-")

    def test_kundennummern_bleiben_eindeutig(self, client):
        """Import und Maske dürfen nicht dieselbe Nummer vergeben."""
        client.post("/api/v1/sales/customers", json={"name": "Erster", "typ": "GEWERBE"})

        tpl = client.get("/api/v1/imports/template/customers").content
        wb = load_workbook(io.BytesIO(tpl))
        ws = wb["Daten"]
        header = [str(c.value or "") for c in ws[1]]
        for name in ("Zweiter", "Dritter"):
            zeile = [None] * len(header)
            zeile[header.index("name *")] = name
            zeile[header.index("typ *")] = "GEWERBE"
            ws.append(zeile)
        buf = io.BytesIO()
        wb.save(buf)
        _upload(client, "customers", buf.getvalue())

        client.post("/api/v1/sales/customers", json={"name": "Vierter", "typ": "GEWERBE"})

        nummern = [k["customer_number"] for k in client.get("/api/v1/sales/customers").json()["items"]]
        assert all(nummern), f"Kunde ohne Nummer: {nummern}"
        assert len(set(nummern)) == len(nummern), f"Doppelte Kundennummern: {nummern}"

    def _kunde_mit_zusatz(self, client):
        """RATIONAL AG mit Adresszusatz 'Werk 2 - Tor 210' (Gernots Fallbeispiel).

        Bindestrich statt Halbgeviertstrich: ReportLab kodiert '–' als WinAnsi
        0x96, das überlebt die Textextraktion im Test nicht — geprüft wird der
        Datenfluss, nicht die Zeichenkodierung.
        """
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()
        r = client.post(f"/api/v1/sales/customers/{kunde['id']}/addresses", json={
            "address_type": "BOTH",
            "strasse": "Siegfried-Meister-Strasse", "hausnummer": "1",
            "adresszusatz": "Werk 2 - Tor 210",
            "plz": "86899", "ort": "Landsberg", "is_default": True,
        })
        assert r.status_code == 201, r.text
        assert r.json()["adresszusatz"] == "Werk 2 - Tor 210"
        return kunde

    def _bestellung(self, client, kunde_id):
        r = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde_id,
            "requested_delivery_date": date.today().isoformat(),
            "lines": [{"product_name": "Kresse", "quantity": 2, "unit": "STK",
                       "unit_price": 3.0, "tax_rate": "REDUZIERT"}],
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def test_adresszusatz_steht_in_der_rechnungsanschrift(self, client):
        """Regression: die Anschrift auf der Rechnung führt den Zusatz."""
        kunde = self._kunde_mit_zusatz(client)
        order = self._bestellung(client, kunde["id"])
        inv = client.post(f"/api/v1/invoices/from-order/{order['id']}")
        assert inv.status_code in (200, 201), inv.text
        pdf = client.get(f"/api/v1/invoices/{inv.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text
        assert "Tor 210" in pdf_text(pdf.content)

    def test_adresszusatz_landet_im_lieferadress_snapshot(self, client):
        """Die Bestellung friert die Lieferadresse ein — ohne den Zusatz ist er
        auf jedem daraus erzeugten Beleg verloren."""
        kunde = self._kunde_mit_zusatz(client)
        order = self._bestellung(client, kunde["id"])
        assert order["delivery_address"], "Lieferadresse nicht übernommen"
        assert order["delivery_address"].get("adresszusatz") == "Werk 2 - Tor 210"

    def test_adresszusatz_steht_auf_dem_lieferschein(self, client):
        """Ohne 'Werk 2 - Tor 210' findet die Spedition das Tor nicht."""
        kunde = self._kunde_mit_zusatz(client)
        order = self._bestellung(client, kunde["id"])
        note = client.post(f"/api/v1/sales/orders/{order['id']}/delivery-notes", json={})
        assert note.status_code == 201, note.text
        pdf = client.get(f"/api/v1/sales/delivery-notes/{note.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text
        assert "Tor 210" in pdf_text(pdf.content)


class TestMehrwertsteuerRundung:
    """Punkt 3.c.iii: 'Netto: 242,84 € müssten bei 7% MWSt 17,00€ sein und
    nicht wie ausgewiesen 17,03€'.

    § 16 Abs. 1 UStG: die Steuer wird auf die Summe der Bemessungsgrundlagen
    je Steuersatz berechnet — also einmal am Ende gerundet. Wurde stattdessen
    je Position gerundet und dann summiert, wanderten 13 × 0,0024 € Rundungs-
    rest in die Steuer: 17,03 € statt 17,00 €.
    """

    #: Gernots Fall: 13 Positionen à 18,68 € netto = 242,84 €
    POSITIONEN = 13
    EINZELPREIS = Decimal("18.68")

    def _order(self, client, show_prices=False):
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
            "show_prices_on_delivery_note": show_prices,
        }).json()
        r = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde["id"],
            "requested_delivery_date": date.today().isoformat(),
            "lines": [
                {"product_name": f"Sorte {i + 1}", "quantity": 1, "unit": "STK",
                 "unit_price": float(self.EINZELPREIS), "tax_rate": "REDUZIERT"}
                for i in range(self.POSITIONEN)
            ],
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def test_bestellsummen_runden_die_steuer_nur_einmal(self, client):
        order = self._order(client)
        assert Decimal(str(order["total_net"])) == Decimal("242.84")
        assert Decimal(str(order["total_vat"])) == Decimal("17.00")
        assert Decimal(str(order["total_gross"])) == Decimal("259.84")

    def test_lieferschein_mit_preisen_weist_17_00_aus(self, client):
        order = self._order(client, show_prices=True)
        note = client.post(f"/api/v1/sales/orders/{order['id']}/delivery-notes", json={})
        assert note.status_code == 201, note.text
        pdf = client.get(f"/api/v1/sales/delivery-notes/{note.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text
        text = pdf_text(pdf.content)
        assert "17.00" in text, "USt-Summe auf dem Lieferschein falsch gerundet"
        assert "17.03" not in text

    def test_gemischte_steuersaetze_bleiben_getrennt(self, client):
        """7 % und 19 % dürfen nicht in einen Topf — sonst stimmt keiner."""
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "Mischbeleg", "typ": "GEWERBE",
        }).json()
        r = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde["id"],
            "requested_delivery_date": date.today().isoformat(),
            "lines": [
                {"product_name": "Kresse", "quantity": 13, "unit": "STK",
                 "unit_price": 18.68, "tax_rate": "REDUZIERT"},
                {"product_name": "Pfandkiste", "quantity": 2, "unit": "STK",
                 "unit_price": 10.00, "tax_rate": "STANDARD"},
            ],
        })
        assert r.status_code in (200, 201), r.text
        order = r.json()
        # 242,84 × 7 % = 17,00 | 20,00 × 19 % = 3,80
        assert Decimal(str(order["total_net"])) == Decimal("262.84")
        assert Decimal(str(order["total_vat"])) == Decimal("20.80")


class TestChargeInVerwendung:
    """Punkt 5c: eine Charge nachträglich auf 'in Verwendung' setzen.

    'In Produktion ab' war nur beim Wareneingang erfassbar. Wer das Datum
    dort vergaß, hatte keinen Weg mehr, die Charge als angebrochen zu
    kennzeichnen.
    """

    def _charge(self, client, sample_seed, sample_location, batch="SB-2026-001"):
        r = client.post("/api/v1/inventory/seeds/receive", params={
            "seed_id": sample_seed["id"], "batch_number": batch,
            "quantity": 5, "unit": "KG", "location_id": sample_location["id"],
        })
        assert r.status_code == 201, r.text
        return r.json()

    def test_bestand_liefert_in_produktion_ab(self, client, sample_seed, sample_location):
        """Ohne den Wert in der Antwort kann die Maske ihn nicht anzeigen."""
        bestand = self._charge(client, sample_seed, sample_location)
        assert "in_production_at" in bestand
        assert bestand["in_production_at"] is None  # frisch eingelagert

    def test_in_produktion_ab_nachtraeglich_setzen(self, client, sample_seed, sample_location):
        bestand = self._charge(client, sample_seed, sample_location, "SB-2026-002")
        heute = date.today().isoformat()

        r = client.patch(f"/api/v1/inventory/seeds/{bestand['id']}",
                         json={"in_production_at": heute})

        assert r.status_code == 200, r.text
        assert r.json()["in_production_at"] == heute

        # und die Traceability-Charge zieht mit (Rückverfolgbarkeit)
        batches = client.get(f"/api/v1/seeds/{sample_seed['id']}/batches").json()
        charge = next(b for b in batches if b["charge_nummer"] == "SB-2026-002")
        assert charge["in_production_at"] == heute

    def test_liste_zeigt_in_produktion_ab(self, client, sample_seed, sample_location):
        bestand = self._charge(client, sample_seed, sample_location, "SB-2026-003")
        heute = date.today().isoformat()
        client.patch(f"/api/v1/inventory/seeds/{bestand['id']}", json={"in_production_at": heute})

        items = client.get("/api/v1/inventory/seeds").json()
        items = items["items"] if isinstance(items, dict) else items
        eintrag = next(i for i in items if i["batch_number"] == "SB-2026-003")
        assert eintrag["in_production_at"] == heute


class TestSorteAufBelegen:
    """Punkt 3.b.i: 'bei sortenreinen Produkten soll die Sorte auf den Belegen
    stehen'.

    Der Artikelname trägt sie nicht ("Sonnenblume 100 g Schale" sagt nichts
    über 'Black Oil'). Die Belegzeile führte bisher nur Bundle-Inhalt und
    EAN als Zusatz — die Sorte fehlte, obwohl sie am Saatgut hängt.
    """

    def _saatgut(self, client, name="Sonnenblume", sorte="Black Oil"):
        r = client.post("/api/v1/seeds", json={
            "name": name, "sorte": sorte,
            "keimdauer_tage": 2, "wachstumsdauer_tage": 8,
            "erntefenster_min_tage": 9, "erntefenster_optimal_tage": 11,
            "erntefenster_max_tage": 14, "ertrag_gramm_pro_tray": 350,
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def _produkt(self, client, sku="MG-SONNE-100", **extra):
        r = client.post("/api/v1/products", json={
            "sku": sku, "name": "Sonnenblume 100 g Schale",
            "category": "MICROGREEN", "base_price": 3.90, **extra,
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def _beleg_text(self, client, produkt):
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()
        order = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde["id"],
            "requested_delivery_date": date.today().isoformat(),
            "lines": [{"product_id": produkt["id"], "product_name": produkt["name"],
                       "quantity": 2, "unit": "STK", "unit_price": 3.90,
                       "tax_rate": "REDUZIERT"}],
        })
        assert order.status_code in (200, 201), order.text
        note = client.post(f"/api/v1/sales/orders/{order.json()['id']}/delivery-notes", json={})
        assert note.status_code == 201, note.text
        pdf = client.get(f"/api/v1/sales/delivery-notes/{note.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text
        return pdf_text(pdf.content)

    def test_lieferschein_nennt_die_sorte_des_saatguts(self, client, base_unit):
        """Sorte hängt am verknüpften Saatgut, nicht am Artikelnamen."""
        saatgut = self._saatgut(client)
        produkt = self._produkt(client, seed_id=saatgut["id"])

        assert "Black Oil" in self._beleg_text(client, produkt)

    def test_sorte_am_produkt_hat_vorrang(self, client, base_unit):
        """Weicht die verkaufte Sorte vom Saatgut-Stammsatz ab, zählt das
        Produktfeld — sonst steht die falsche Sorte auf dem Beleg."""
        saatgut = self._saatgut(client)
        produkt = self._produkt(client, sku="MG-SONNE-200",
                                seed_id=saatgut["id"], seed_variety="Peredovik")

        text = self._beleg_text(client, produkt)
        assert "Peredovik" in text
        assert "Black Oil" not in text

    def test_rechnung_nennt_die_sorte(self, client, base_unit):
        """Die Rechnung ist der Beleg, der beim Kunden im Ordner landet."""
        saatgut = self._saatgut(client, "Erbse", "Gruenschnitt")
        produkt = self._produkt(client, sku="MG-ERBSE-100", seed_id=saatgut["id"])
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()
        order = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde["id"],
            "requested_delivery_date": date.today().isoformat(),
            "lines": [{"product_id": produkt["id"], "product_name": produkt["name"],
                       "quantity": 2, "unit": "STK", "unit_price": 3.90,
                       "tax_rate": "REDUZIERT"}],
        })
        assert order.status_code in (200, 201), order.text
        inv = client.post(f"/api/v1/invoices/from-order/{order.json()['id']}")
        assert inv.status_code in (200, 201), inv.text
        pdf = client.get(f"/api/v1/invoices/{inv.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text

        assert "Gruenschnitt" in pdf_text(pdf.content)

    def test_produkt_import_uebernimmt_die_sorte(self, client, base_unit):
        """Gernot pflegt seinen Katalog per Excel — ohne Spalte keine Sorte."""
        from openpyxl import Workbook

        tpl = client.get("/api/v1/imports/template/products").content
        header = [c.value for c in load_workbook(io.BytesIO(tpl))["Beispiel"][1]]
        assert any(str(h).startswith("sorte") for h in header), "Spalte 'sorte' fehlt"

        wb = Workbook()
        ws = wb.active
        ws.title = "Daten"
        ws.append(header)
        ws.append([c.value for c in load_workbook(io.BytesIO(tpl))["Beispiel"][2]])
        buf = io.BytesIO()
        wb.save(buf)

        r = client.post("/api/v1/imports/products",
                        files={"file": ("products.xlsx", buf.getvalue(), XLSX_MIME)})
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 1, r.json()

        katalog = client.get("/api/v1/products").json()
        katalog = katalog["items"] if isinstance(katalog, dict) else katalog
        produkt = next(p for p in katalog if p["sku"] == "MG-RUC-100")
        assert produkt["seed_variety"] == "Coltivata"

    def test_produkt_ohne_sorte_bleibt_unveraendert(self, client, base_unit):
        """Kein Saatgut, keine Sorte — die Zeile darf kein leeres 'Sorte:'
        anhängen."""
        produkt = self._produkt(client, sku="MG-SONNE-300")

        assert "Sorte:" not in self._beleg_text(client, produkt)


class TestPfandartikel:
    """Punkt 4: 'Pfandwesen fehlt' — Pfandtrays und Pfandkisten mit 19 % MwSt.

    Das Datenmodell kannte Pfand längst (Product.is_deposit, InvoiceLine.
    is_deposit, Invoice.total_deposit), nur führte kein Weg dorthin: weder
    Schema noch Maske noch Import kannten die Felder. Damit ließ sich schlicht
    kein Pfandartikel anlegen.
    """

    def _pfandartikel(self, client, sku="PFAND-KISTE", **extra):
        payload = {
            "sku": sku, "name": "Pfandkiste E2", "category": "PACKAGING",
            "base_price": 4.50, "is_deposit": True, "deposit_value": 4.50,
        }
        payload.update(extra)
        r = client.post("/api/v1/products", json=payload)
        assert r.status_code in (200, 201), r.text
        return r.json()

    def test_pfandartikel_anlegen(self, client, base_unit):
        artikel = self._pfandartikel(client)
        assert artikel["is_deposit"] is True
        assert Decimal(str(artikel["deposit_value"])) == Decimal("4.50")

    def test_pfand_ist_standardsteuersatz(self, client, base_unit):
        """Pfand auf Mehrweggebinde ist ein eigener Umsatz zu 19 % — der
        Lebensmittelsatz von 7 % gilt dafür nicht."""
        artikel = self._pfandartikel(client, sku="PFAND-TRAY")
        assert artikel["tax_rate"] == "STANDARD"

    def test_expliziter_steuersatz_bleibt_erhalten(self, client, base_unit):
        """Der Vorschlag darf eine bewusste Angabe nicht überschreiben."""
        artikel = self._pfandartikel(client, sku="PFAND-SONDER", tax_rate="REDUZIERT")
        assert artikel["tax_rate"] == "REDUZIERT"

    def test_pfandkennzeichen_nachtraeglich_setzen(self, client, base_unit):
        r = client.post("/api/v1/products", json={
            "sku": "KISTE-NEU", "name": "Mehrwegkiste", "category": "PACKAGING",
            "base_price": 3.00,
        })
        assert r.status_code in (200, 201), r.text

        r = client.patch(f"/api/v1/products/{r.json()['id']}",
                         json={"is_deposit": True, "deposit_value": 3.00})

        assert r.status_code == 200, r.text
        assert r.json()["is_deposit"] is True
        assert Decimal(str(r.json()["deposit_value"])) == Decimal("3.00")

    def test_rechnung_weist_pfand_gesondert_aus(self, client, base_unit):
        """Gernot muss sehen, wie viel Pfand in der Rechnung steckt — das Geld
        gehört ihm nicht, es kommt mit dem Gebinde zurück."""
        pfand = self._pfandartikel(client, sku="PFAND-RECHNUNG")
        kunde = client.post("/api/v1/sales/customers", json={
            "name": "RATIONAL AG", "typ": "GEWERBE",
        }).json()
        order = client.post("/api/v1/sales/orders", json={
            "customer_id": kunde["id"],
            "requested_delivery_date": date.today().isoformat(),
            "lines": [
                {"product_name": "Kresse", "quantity": 10, "unit": "STK",
                 "unit_price": 3.00, "tax_rate": "REDUZIERT"},
                {"product_id": pfand["id"], "product_name": pfand["name"],
                 "quantity": 2, "unit": "STK", "unit_price": 4.50,
                 "tax_rate": "STANDARD"},
            ],
        })
        assert order.status_code in (200, 201), order.text
        inv = client.post(f"/api/v1/invoices/from-order/{order.json()['id']}")
        assert inv.status_code in (200, 201), inv.text

        # 2 × 4,50 € netto + 19 % = 10,71 € brutto Pfand
        assert Decimal(str(inv.json()["total_deposit"])) == Decimal("10.71")

        pdf = client.get(f"/api/v1/invoices/{inv.json()['id']}/pdf")
        assert pdf.status_code == 200, pdf.text
        text = pdf_text(pdf.content)
        assert "Pfand" in text, "Pfandzeile fehlt im Summenblock"
        assert "10.71" in text

    def test_import_legt_pfandartikel_an(self, client, base_unit):
        """Der Katalog kommt per Excel — ohne Spalten kein Pfand."""
        from openpyxl import Workbook

        tpl = client.get("/api/v1/imports/template/products").content
        header = [c.value for c in load_workbook(io.BytesIO(tpl))["Daten"][1]]
        spalten = [str(h).rstrip(" *") for h in header]
        assert "pfand" in spalten, f"Spalte 'pfand' fehlt: {spalten}"
        assert "pfandwert" in spalten, f"Spalte 'pfandwert' fehlt: {spalten}"

        zeile = [""] * len(header)
        werte = {"sku": "PFAND-IMP", "name": "Pfandtray", "category": "PACKAGING",
                 "tax_rate": "STANDARD", "pfand": "ja", "pfandwert": "1.50"}
        for spalte, wert in werte.items():
            zeile[spalten.index(spalte)] = wert

        wb = Workbook()
        ws = wb.active
        ws.title = "Daten"
        ws.append(header)
        ws.append(zeile)
        buf = io.BytesIO()
        wb.save(buf)

        r = client.post("/api/v1/imports/products",
                        files={"file": ("products.xlsx", buf.getvalue(), XLSX_MIME)})
        assert r.status_code == 200, r.text
        assert r.json()["created"] == 1, r.json()

        katalog = client.get("/api/v1/products").json()
        katalog = katalog["items"] if isinstance(katalog, dict) else katalog
        artikel = next(p for p in katalog if p["sku"] == "PFAND-IMP")
        assert artikel["is_deposit"] is True
        assert Decimal(str(artikel["deposit_value"])) == Decimal("1.50")


class TestAussaatEtiketten:
    """Punkt 9b: Etikettenbogen für die Aussaat eines Tages.

    Bisher gab es nur ein Einzel-Label je Charge auf 62 × 100 mm Thermopapier.
    Gernot klebt die Etiketten aber auf jedes Tray und druckt sie auf
    Avery-Zweckform-Bögen (48,5 × 16,9 mm, 64 Stück je A4) — für alle Sorten,
    die an einem Tag ausgesät werden, in einem Rutsch.
    """

    ROUTE = "/api/v1/production/labels/grow-batches"

    def _aussaat(self, client, seed_id, charge, trays, tag=None):
        batch = client.post("/api/v1/seeds/batches", json={
            "seed_id": seed_id, "charge_nummer": charge, "menge_gramm": 5000,
        })
        assert batch.status_code in (200, 201), batch.text
        r = client.post("/api/v1/production/grow-batches", json={
            "seed_batch_id": batch.json()["id"],
            "tray_anzahl": trays,
            "aussaat_datum": (tag or date.today()).isoformat(),
        })
        assert r.status_code == 201, r.text
        return r.json()

    def _sorte(self, client, name, sorte):
        r = client.post("/api/v1/seeds", json={
            "name": name, "sorte": sorte,
            "keimdauer_tage": 2, "wachstumsdauer_tage": 8,
            "erntefenster_min_tage": 9, "erntefenster_optimal_tage": 11,
            "erntefenster_max_tage": 14, "ertrag_gramm_pro_tray": 350,
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def test_bogen_nennt_sorte_datum_und_charge(self, client, sample_seed):
        """Ohne Chargennummer ist das Tray nicht rückverfolgbar."""
        self._aussaat(client, sample_seed["id"], "SB-0901", 2)

        r = client.get(self.ROUTE, params={"datum": date.today().isoformat()})

        assert r.status_code == 200, r.text
        assert r.headers["content-type"] == "application/pdf"
        text = pdf_text(r.content)
        assert "Sonnenblume" in text
        assert date.today().strftime("%d.%m.%Y") in text
        assert "SB-0901" in text

    def test_ein_etikett_je_tray(self, client, sample_seed):
        """Jedes Tray bekommt sein eigenes Etikett — durchnummeriert."""
        self._aussaat(client, sample_seed["id"], "SB-0902", 3)

        text = pdf_text(client.get(self.ROUTE, params={
            "datum": date.today().isoformat()}).content)

        assert "1/3" in text
        assert "2/3" in text
        assert "3/3" in text

    def test_alle_sorten_des_tages_auf_einem_bogen(self, client, sample_seed):
        erbse = self._sorte(client, "Erbse", "Zuckererbse")
        self._aussaat(client, sample_seed["id"], "SB-0903", 1)
        self._aussaat(client, erbse["id"], "SB-0904", 1)

        text = pdf_text(client.get(self.ROUTE, params={
            "datum": date.today().isoformat()}).content)

        assert "Sonnenblume" in text
        assert "Erbse" in text

    def test_andere_tage_bleiben_draussen(self, client, sample_seed):
        gestern = date.today() - timedelta(days=1)
        self._aussaat(client, sample_seed["id"], "SB-0905", 1, tag=gestern)
        erbse = self._sorte(client, "Erbse", "Zuckererbse")
        self._aussaat(client, erbse["id"], "SB-0906", 1)

        text = pdf_text(client.get(self.ROUTE, params={
            "datum": date.today().isoformat()}).content)

        assert "Erbse" in text
        assert "SB-0905" not in text

    def test_ohne_aussaat_klare_meldung(self, client):
        """Ein leerer Bogen im Drucker hilft niemandem."""
        r = client.get(self.ROUTE, params={
            "datum": (date.today() + timedelta(days=30)).isoformat()})

        assert r.status_code == 404
        assert "Aussaat" in r.json()["detail"]

    def test_avery_bogen_ist_a4(self, client, sample_seed):
        """48,5 × 16,9 mm liegen auf A4 — 64 Etiketten je Bogen."""
        self._aussaat(client, sample_seed["id"], "SB-0907", 1)

        raw = client.get(self.ROUTE, params={
            "datum": date.today().isoformat()}).content

        # ReportLab schreibt die Seitengröße in Punkt: A4 = 595 × 842
        assert b"/MediaBox [ 0 0 595.2756 841.8898 ]" in raw

    def test_kleines_format_ein_etikett_je_seite(self, client, sample_seed):
        """45 × 25 mm läuft auf dem Rollendrucker — ein Etikett je Seite."""
        self._aussaat(client, sample_seed["id"], "SB-0908", 2)

        r = client.get(self.ROUTE, params={
            "datum": date.today().isoformat(), "format": "45x25"})

        assert r.status_code == 200, r.text
        # 45 mm = 127,559 pt | 25 mm = 70,866 pt
        assert b"/MediaBox [ 0 0 127.5591 70.866" in r.content
        assert r.content.count(b"/MediaBox") == 2  # zwei Trays, zwei Seiten


class TestSaatgutMischung:
    """Punkt 5b: 'Brotzeitmix' aus mehreren Sorten als eigene Sorte.

    Der Mix wird nicht eingekauft, sondern beim Aussäen gemischt. Bisher ließ
    er sich nur als normale Sorte anlegen — ohne Bestandsabzug bei den
    Ausgangssorten und ohne Spur, welche Chargen darin stecken.
    """

    def _sorte(self, client, name, **extra):
        r = client.post("/api/v1/seeds", json={
            "name": name, "keimdauer_tage": 2, "wachstumsdauer_tage": 8,
            "erntefenster_min_tage": 9, "erntefenster_optimal_tage": 11,
            "erntefenster_max_tage": 14, "ertrag_gramm_pro_tray": 350, **extra,
        })
        assert r.status_code in (200, 201), r.text
        return r.json()

    def _bestand(self, client, sample_location, seed, charge, gramm):
        r = client.post("/api/v1/inventory/seeds/receive", params={
            "seed_id": seed["id"], "batch_number": charge,
            "quantity": gramm, "unit": "G", "location_id": sample_location["id"],
        })
        assert r.status_code == 201, r.text
        return r.json()

    def _mix(self, client, sample_location, mengen=(20, 30), bestand=1000):
        """Mix aus zwei Sorten; gibt (mix, [komponenten]) zurück."""
        komponenten = []
        for nr, gramm in enumerate(mengen, start=1):
            sorte = self._sorte(client, f"Komponente {nr}")
            self._bestand(client, sample_location, sorte, f"KOMP-{nr}", bestand)
            komponenten.append((sorte, gramm))
        mix = self._sorte(client, "Brotzeitmix", is_mix=True, mix_components=[
            {"seed_id": s["id"], "gramm_pro_tray": g} for s, g in komponenten
        ])
        return mix, komponenten

    def _menge_kg(self, client, charge):
        items = client.get("/api/v1/inventory/seeds").json()
        items = items["items"] if isinstance(items, dict) else items
        eintrag = next(i for i in items if i["batch_number"] == charge)
        return Decimal(str(eintrag["current_quantity_kg"]))

    def test_mischsorte_anlegen(self, client, sample_location):
        mix, komponenten = self._mix(client, sample_location)

        assert mix["is_mix"] is True
        assert len(mix["mix_components"]) == 2
        namen = {k["seed_name"] for k in mix["mix_components"]}
        assert namen == {"Komponente 1", "Komponente 2"}

    def test_aussaat_zieht_von_den_ausgangssorten_ab(self, client, sample_location):
        """4 Kisten × 20 g + 4 × 30 g = 80 g und 120 g weniger im Lager."""
        mix, _ = self._mix(client, sample_location)

        r = client.post("/api/v1/production/grow-batches", json={
            "seed_id": mix["id"], "tray_anzahl": 4,
            "aussaat_datum": date.today().isoformat(),
        })

        assert r.status_code == 201, r.text
        assert self._menge_kg(client, "KOMP-1") == Decimal("0.920")
        assert self._menge_kg(client, "KOMP-2") == Decimal("0.880")

    def test_chargennummern_werden_festgehalten(self, client, sample_location):
        """Rückverfolgbarkeit: welche Charge steckt in welcher Mischung?"""
        mix, _ = self._mix(client, sample_location)
        charge = client.post("/api/v1/production/grow-batches", json={
            "seed_id": mix["id"], "tray_anzahl": 4,
            "aussaat_datum": date.today().isoformat(),
        }).json()

        r = client.get(f"/api/v1/seeds/batches/{charge['seed_batch_id']}/components")

        assert r.status_code == 200, r.text
        verwendet = {k["charge_nummer"]: Decimal(str(k["menge_gramm"])) for k in r.json()}
        assert verwendet == {"KOMP-1": Decimal("80.00"), "KOMP-2": Decimal("120.00")}

    def test_mischcharge_traegt_den_mixnamen(self, client, sample_location):
        mix, _ = self._mix(client, sample_location)
        charge = client.post("/api/v1/production/grow-batches", json={
            "seed_id": mix["id"], "tray_anzahl": 2,
            "aussaat_datum": date.today().isoformat(),
        }).json()

        assert charge["seed_name"] == "Brotzeitmix"
        batches = client.get(f"/api/v1/seeds/{mix['id']}/batches").json()
        assert batches[0]["charge_nummer"].startswith("MIX-")

    def test_zu_wenig_saatgut_bricht_ganz_ab(self, client, sample_location):
        """Halb abgezogen wäre schlimmer als gar nicht: Bestand bleibt stehen."""
        mix, _ = self._mix(client, sample_location, mengen=(20, 30), bestand=100)

        r = client.post("/api/v1/production/grow-batches", json={
            "seed_id": mix["id"], "tray_anzahl": 5,
            "aussaat_datum": date.today().isoformat(),
        })

        assert r.status_code == 400, r.text
        assert "Komponente 2" in r.json()["detail"]
        assert self._menge_kg(client, "KOMP-1") == Decimal("0.100")
        assert self._menge_kg(client, "KOMP-2") == Decimal("0.100")

    def test_mix_ohne_komponenten_wird_abgelehnt(self, client, sample_location):
        leer = self._sorte(client, "Leerer Mix", is_mix=True)

        r = client.post("/api/v1/production/grow-batches", json={
            "seed_id": leer["id"], "tray_anzahl": 1,
            "aussaat_datum": date.today().isoformat(),
        })

        assert r.status_code == 400
        assert "Komponente" in r.json()["detail"]

    def test_normale_aussaat_bleibt_unveraendert(self, client, sample_seed):
        """Regressionswächter: der bestehende Weg über die Charge muss halten."""
        batch = client.post("/api/v1/seeds/batches", json={
            "seed_id": sample_seed["id"], "charge_nummer": "SB-MIX-REG",
            "menge_gramm": 5000,
        }).json()

        r = client.post("/api/v1/production/grow-batches", json={
            "seed_batch_id": batch["id"], "tray_anzahl": 3,
            "aussaat_datum": date.today().isoformat(),
        })

        assert r.status_code == 201, r.text
        assert r.json()["seed_name"] == "Sonnenblume"
        components = client.get(f"/api/v1/seeds/batches/{batch['id']}/components")
        assert components.json() == []


class TestSaatgutEntnahme:
    """Lager → Saatgut entnehmen.

    Beim Bauen der Mischung aufgefallen: der Endpunkt packte die Antwort des
    Services falsch aus und lief in einen 500er — die Handentnahme im Lager
    war damit unbenutzbar.
    """

    def test_entnahme_bucht_ab(self, client, sample_seed, sample_location):
        bestand = client.post("/api/v1/inventory/seeds/receive", params={
            "seed_id": sample_seed["id"], "batch_number": "ENT-1",
            "quantity": 1000, "unit": "G", "location_id": sample_location["id"],
        }).json()

        r = client.post(
            f"/api/v1/inventory/seeds/{bestand['id']}/consume",
            params={"quantity": 0.25},
        )

        assert r.status_code == 200, r.text
        assert Decimal(str(r.json()["inventory"]["current_quantity_kg"])) == Decimal("0.750")
        # Abgang wird negativ gebucht
        assert Decimal(str(r.json()["movement"]["quantity"])) == Decimal("-0.250")

    def test_mehr_als_vorhanden_wird_abgelehnt(self, client, sample_seed, sample_location):
        bestand = client.post("/api/v1/inventory/seeds/receive", params={
            "seed_id": sample_seed["id"], "batch_number": "ENT-2",
            "quantity": 100, "unit": "G", "location_id": sample_location["id"],
        }).json()

        r = client.post(
            f"/api/v1/inventory/seeds/{bestand['id']}/consume",
            params={"quantity": 1},
        )

        assert r.status_code == 400, r.text
